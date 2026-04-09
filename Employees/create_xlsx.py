import csv
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook


BASE_DIR = Path(__file__).resolve().parent
DATA_FILE = BASE_DIR / "data.csv"
OUTPUT_FILE = BASE_DIR / "employees.xlsx"

SHEET_TITLES = {
    "all": "Усі працівники",
    "younger_18": "До 18",
    "18-45": "18-45",
    "45-70": "45-70",
    "older_70": "Понад 70",
}


def calculate_age(birth_date):
    today = datetime.today().date()
    return today.year - birth_date.year - (
        (today.month, today.day) < (birth_date.month, birth_date.day)
    )


try:
    with DATA_FILE.open(encoding="utf-8") as file:
        reader = list(csv.reader(file))
except OSError:
    print("Помилка відкриття CSV")
    raise SystemExit(1)

wb = Workbook()
sheets = {
    "all": wb.active,
    "younger_18": wb.create_sheet(title=SHEET_TITLES["younger_18"]),
    "18-45": wb.create_sheet(title=SHEET_TITLES["18-45"]),
    "45-70": wb.create_sheet(title=SHEET_TITLES["45-70"]),
    "older_70": wb.create_sheet(title=SHEET_TITLES["older_70"]),
}

sheets["all"].title = SHEET_TITLES["all"]

for sheet in sheets.values():
    sheet.append(["№", "Прізвище", "Ім'я", "По батькові", "Дата народження", "Вік"])

counter = 1

for row in reader[1:]:
    birth_date = datetime.strptime(row[4], "%Y-%m-%d").date()
    age = calculate_age(birth_date)

    record = [counter, row[0], row[1], row[2], row[4], age]
    sheets["all"].append(record)

    if age < 18:
        sheets["younger_18"].append(record)
    elif age <= 45:
        sheets["18-45"].append(record)
    elif age <= 70:
        sheets["45-70"].append(record)
    else:
        sheets["older_70"].append(record)

    counter += 1

try:
    wb.save(OUTPUT_FILE)
    print("Ok")
except OSError:
    print("Не вдалося створити XLSX")
